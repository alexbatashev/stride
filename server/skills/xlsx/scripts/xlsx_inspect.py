import argparse
import json

from openpyxl import load_workbook


FORMULA_ERRORS = {
    "#NULL!",
    "#DIV/0!",
    "#VALUE!",
    "#REF!",
    "#NAME?",
    "#NUM!",
    "#N/A",
    "#GETTING_DATA",
    "#SPILL!",
    "#CALC!",
    "#FIELD!",
    "#BLOCKED!",
    "#UNKNOWN!",
    "#CONNECT!",
    "#BUSY!",
    "#PYTHON!",
}


def calculation_settings(workbook):
    calculation = workbook.calculation
    return {
        name: getattr(calculation, name)
        for name in (
            "calcMode",
            "calcId",
            "fullCalcOnLoad",
            "forceFullCalc",
            "iterate",
            "iterateCount",
            "iterateDelta",
        )
        if getattr(calculation, name) is not None
    }


def inspect_workbook(path, detail_limit):
    formulas_book = load_workbook(
        path, data_only=False, read_only=False, keep_links=True
    )
    values_book = load_workbook(path, data_only=True, read_only=False, keep_links=True)
    formula_details = []
    formula_count = 0
    missing_cache = 0
    cached_errors = []
    sheets = []

    for sheet in formulas_book.worksheets:
        value_sheet = values_book[sheet.title]
        sheet_formulas = 0
        sheet_missing = 0
        sheet_errors = []
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type != "f":
                    continue
                formula_count += 1
                sheet_formulas += 1
                cached = value_sheet[cell.coordinate].value
                if cached is None:
                    missing_cache += 1
                    sheet_missing += 1
                if isinstance(cached, str) and cached in FORMULA_ERRORS:
                    location = f"{sheet.title}!{cell.coordinate}"
                    cached_errors.append(location)
                    sheet_errors.append(location)
                if len(formula_details) < detail_limit:
                    formula_details.append(
                        {
                            "cell": f"{sheet.title}!{cell.coordinate}",
                            "formula": cell.value,
                            "cached": cached,
                        }
                    )
        sheets.append(
            {
                "name": sheet.title,
                "state": sheet.sheet_state,
                "rows": sheet.max_row,
                "columns": sheet.max_column,
                "merged_ranges": len(sheet.merged_cells.ranges),
                "tables": sorted(sheet.tables),
                "formulas": sheet_formulas,
                "missing_formula_cache": sheet_missing,
                "cached_formula_errors": sheet_errors,
                "freeze_panes": str(sheet.freeze_panes) if sheet.freeze_panes else None,
            }
        )

    return {
        "path": path,
        "sheets": sheets,
        "defined_names": sorted(formulas_book.defined_names),
        "external_links": len(formulas_book._external_links),
        "calculation": calculation_settings(formulas_book),
        "formula_summary": {
            "count": formula_count,
            "missing_cache": missing_cache,
            "cached_errors": cached_errors,
        },
        "formula_details": formula_details,
    }


def main():
    parser = argparse.ArgumentParser(description="Inspect an XLSX and print JSON.")
    parser.add_argument("input")
    parser.add_argument("--details", type=int, default=100)
    args = parser.parse_args()
    if args.details < 0:
        raise SystemExit("--details must be non-negative")
    print(
        json.dumps(
            inspect_workbook(args.input, args.details),
            indent=2,
            ensure_ascii=False,
            default=str,
        )
    )


if __name__ == "__main__":
    main()
